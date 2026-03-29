import json
import os
import uuid

from openpyxl import Workbook, load_workbook


class Database:
    HEADERS = ["UUID", "Date", "Store", "Item", "Price"]

    def __init__(self):
        self.base_dir = "../data/users_data"
        self.users_file = os.path.join(self.base_dir, "users.json")
        self._ensure_users_file()

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------

    def _ensure_users_file(self):
        os.makedirs(self.base_dir, exist_ok=True)
        if not os.path.exists(self.users_file):
            with open(self.users_file, "w") as f:
                json.dump({}, f)

    def _user_dir(self, user_id: int | str) -> str:
        path = os.path.join(self.base_dir, str(user_id))
        os.makedirs(path, exist_ok=True)
        return path

    def _xlsx_path(self, user_id: int | str) -> str:
        return os.path.join(self._user_dir(user_id), "data.xlsx")

    def _load_or_create_workbook(self, user_id: int | str):
        path = self._xlsx_path(user_id)
        if os.path.exists(path):
            return load_workbook(path)
        wb = Workbook()
        ws = wb.active
        ws.append(self.HEADERS)
        wb.save(path)
        return wb

    # ------------------------------------------------------------------
    # User metadata (users.json)
    # ------------------------------------------------------------------

    def save_user(self, user_id: int | str, user_data: dict):
        """Persist arbitrary user metadata (name, username, etc.)."""
        with open(self.users_file, "r") as f:
            users = json.load(f)
        users[str(user_id)] = user_data
        with open(self.users_file, "w") as f:
            json.dump(users, f, indent=2, ensure_ascii=False)

    def get_user(self, user_id: int | str) -> dict | None:
        with open(self.users_file, "r") as f:
            users = json.load(f)
        return users.get(str(user_id))

    # ------------------------------------------------------------------
    # Receipt data
    # ------------------------------------------------------------------

    def save_receipt(self, user_id: int | str, receipt: dict):
        """
        Append receipt items to the user's Excel file.

        Expected receipt format:
        {
            "date":  "2024-01-01",
            "store": "Silpo",
            "items": [
                {"name": "Milk", "price": 45.0},
                ...
            ]
        }
        """
        wb = self._load_or_create_workbook(user_id)
        ws = wb.active
        receipt_uuid = str(uuid.uuid4())
        for item in receipt["items"]:
            ws.append([
                receipt_uuid,
                receipt.get("date", ""),
                receipt.get("store", ""),
                item["name"],
                item["price"],
            ])
        wb.save(self._xlsx_path(user_id))

    def get_receipts(self, user_id: int | str) -> list[dict] | None:
        """
        Return all rows as a list of dicts, or None if the user has no data.
        """
        path = self._xlsx_path(user_id)
        if not os.path.exists(path):
            return None
        wb = load_workbook(path)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) <= 1:   # only headers
            return []
        headers = rows[0]
        return [dict(zip(headers, row)) for row in rows[1:]]

    def get_xlsx_path(self, user_id: int | str) -> str | None:
        """
        Return the absolute path to the user's .xlsx file,
        or None if it doesn't exist yet. Pass directly to FSInputFile.
        """
        path = self._xlsx_path(user_id)
        return os.path.abspath(path) if os.path.exists(path) else None