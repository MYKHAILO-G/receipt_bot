from openpyxl import Workbook, load_workbook
import json
import os
import uuid

FILE_NAME = "../data/receipts.xlsx"

def save_to_excel(json_data):
  #  data = json.loads(json_data)
    data = json_data
    if not os.path.exists(FILE_NAME):
        wb = Workbook()
        ws = wb.active
        ws.append(["UUID", "Date", "Store", "Item", "Price"])
        wb.save(FILE_NAME)

    wb = load_workbook(FILE_NAME)
    ws = wb.active

    for item in data["items"]:
        ws.append([
            uuid.uuid4(),
            data["date"],
            data["store"],
            item["name"],
            item["price"],
        ])

    wb.save(FILE_NAME)

def get_excel_data():
    if not os.path.exists(FILE_NAME):
        return None
    
    return load_workbook(FILE_NAME)